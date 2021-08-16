import os
import time
import pandas as pd
import threading
import io
import re

from werkzeug.utils import secure_filename
from flask import flash, current_app
from datasource_manager import CSV_HEADERS
from datasource_manager import DATASOURCE_MAPPING
from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from donations_importer import validate_import_sfd
from shifts_importer import validate_import_vs

from constants import RAW_DATA_PATH

SUCCESS_MSG = 'Uploaded Successfully!'
lock = threading.Lock()


def validate_and_arrange_upload(file):
    current_app.logger.info("Start uploading file: " + file.filename)
    filename = secure_filename(file.filename)
    file_extension = filename.rpartition('.')[2]
    determine_upload_type(file, file_extension)


def determine_upload_type(file, file_extension):
    df = None

    if file_extension == 'csv':
        dfs = [pd.read_csv(io.BytesIO(file.stream.read()), encoding='iso-8859-1')]
        file.close()
    else:
        
        match = re.search('donat', file.filename, re.I)

        if match:   # It's a SalesForce Donations file
            validate_import_sfd(file)
            return
        else:
            match = re.search('volunteer', file.filename, re.I)
            if match:    # It's a Volgistics file
                validate_import_vs(file)  
                dfs = excel_to_dataframes(file)  # Also need to run Volgistics through match processing
            else:
                dfs = excel_to_dataframes(file)   #  It's a non-Volgistics, non-Shelterluv XLS? file 


  

    found_sources = 0
    for df in dfs:
        for src_type in CSV_HEADERS:
            if set(CSV_HEADERS[src_type]).issubset(df.columns):
                with lock:
                    found_sources += 1
                    filename = secure_filename(file.filename)
                    now = time.localtime()
                    now_date = time.strftime("%Y-%m-%d--%H-%M-%S", now)
                    current_app.logger.info("  -File: " + filename + " Matches files type: " + src_type)
                    df.to_csv(os.path.join(RAW_DATA_PATH, src_type + '-' + now_date + '.csv'))
                    clean_current_folder(src_type)
                    df.to_csv(os.path.join(RAW_DATA_PATH, src_type + '-' + now_date + '.csv'))
                    current_app.logger.info("  -Uploaded successfully as : " + src_type + '-' + now_date + '.' + file_extension)
                    flash(src_type + " {0} ".format(SUCCESS_MSG), 'info')
    if found_sources == 0:
        current_app.logger.error("\n\n          !!!!!!! No sources found in upload !!!!  \n                Uploaded file " + file.filename + " is probably from wrong report \n          !!!!!!!!!!!")


def excel_to_dataframes(xls):
    df = []
    wb = load_workbook(xls)

    if len(wb.sheetnames) > 1:
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            for sheetname in wb.sheetnames:
                for item in DATASOURCE_MAPPING:
                    if 'sheetname' in DATASOURCE_MAPPING[item]:
                        if DATASOURCE_MAPPING[item]['sheetname'] == sheetname:
                            tmp.seek(0)
                            df.append(pd.read_excel(tmp.read(), sheetname))
    else:
        df.append(pd.read_excel(xls))

    return df


def clean_current_folder(src_type):
    if os.listdir(RAW_DATA_PATH):
        for file_name in os.listdir(RAW_DATA_PATH):
            file_path = os.path.join(RAW_DATA_PATH, file_name)
            file_name_striped = file_path.split('-')[0].split('/')[-1]

            if file_name_striped == src_type:
                current_app.logger.info('File to remove: ' + file_path)
                os.remove(file_path)
                current_app.logger.info("  -Removed file: " + file_name + " from Current files folder")
