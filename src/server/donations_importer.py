import re
from flask.globals import current_app

from openpyxl import load_workbook
from jellyfish import jaro_similarity

from config import  engine

import structlog
logger = structlog.get_logger()

from sqlalchemy import  insert,  Table,  Column, MetaData, exc
from sqlalchemy.dialects.postgresql import Insert
metadata = MetaData()



MINIMUM_SIMILARITY = 0.85  # How good does the table match need to be?

expected_columns =  {'Recurring donor' : 'recurring_donor',        # 'Export XLSX file column name' : 'db column name' 
                    'Opportunity Owner': None ,                    # None means we won't import that column into DB
                    'Account ID 18': None ,
                    'Account Name':  None, 
                    'Primary Contact': 'primary_contact',
                    'Contact ID 18': 'contact_id',
                    'Opportunity ID 18': 'opp_id',   # Should be a unique donation ID but isn't quite
                    'Opportunity Name': None,
                    'Stage': None,
                    'Fiscal Period': None, 
                    'Amount': 'amount', 
                    'Probability (%)': None, 
                    'Age': None,
                    'Close Date': 'close_date', 
                    'Created Date': None, 
                    'Type': 'donation_type', 
                    'Primary Campaign Source': 'primary_campaign_source' ,
                    'Source': None }
                    

def validate_import_sfd(filename, conn):
    """ Validate that the XLSX column names int the file are close enough to expectations that we can trust the data.
        If so, insert the data into the salseforcedonations table. 
    """

    logger.info('---------------------- Loading %s -------------------',  filename.filename)
    wb = load_workbook(filename)   #  ,read_only=True should be faster but gets size incorrect 
    ws = wb.active
    # ws.reset_dimensions()   # Tells openpyxl to ignore what sheet says and check for itself
    ws.calculate_dimension()

    columns = ws.max_column
    if columns > 26:
        # TODO: Handle AA, AB, usw...
        logger.warning("Column count > 26; columns after Z not processed")
        columns = 26

    header = [cell.value for cell in ws[1]]

    min_similarity = 1.0
    min_column = None

    for expected, got in zip(expected_columns.keys(), header):
        jsim = jaro_similarity(expected, got) 
        if jsim < min_similarity :
            min_similarity = jsim
            min_column = expected + ' / ' + got
    
    
    logger.debug("Minimum similarity: %s",  "{:.2}".format(min_similarity) )
    if min_column:
        logger.debug("On expected/got: %s", str(min_column))

    if  min_similarity >= MINIMUM_SIMILARITY :  # Good enough to trust
        
        sfd  = Table("salesforcedonations", metadata, autoload=True, autoload_with=engine)

        seen_header = False  # Used to skip header row

        # Stats for import
        dupes = 0
        other_integrity = 0
        other_exceptions = 0
        row_count = 0
        missing_contact_id = 0

        for row in ws.values:        
            if seen_header: 
                row_count += 1
                if row_count % 1000 == 0:
                    logger.debug("Row: %s",  str(row_count) )
                zrow = dict(zip(expected_columns.values(), row))  
                # zrow is a dict of db_col:value pairs, with at most one key being None (as it overwrote any previous)
                # We need to remove the None item, if it exists
                try:
                    del zrow[None]
                except KeyError:
                    pass 

                #  Cleanup time!  Many older imports have... peculiarities 
                if zrow['amount'] == None:  # We get some with no value, probably user error
                    zrow['amount'] = 0.0    # Setting bad amounts to 0 as per KF

                if zrow['recurring_donor'] == '=FALSE()' :
                    zrow['recurring_donor'] = False

                if zrow['recurring_donor'] == '=TRUE()' :
                    zrow['recurring_donor'] = True

                #  End cleanup 

                if zrow['contact_id'] :  # No point in importing if there's nothing to match
                    # Finally ready to insert row into the table
                    # 

                    stmt = Insert(sfd).values(zrow)

                    skip_dupes = stmt.on_conflict_do_nothing(
                        constraint='uq_donation'
                       )
                    try:
                        result = conn.execute(skip_dupes)
                    except exc.IntegrityError as e:  # Catch-all for several more specific exceptions
                        if  re.match('duplicate key value', str(e.orig) ):
                            dupes += 1
                            pass
                        else:
                            other_integrity += 1
                            logger.error(e)
                    except Exception as e: 
                        other_exceptions += 1
                        logger.error(e)
                 
                else: # Missing contact_id
                    missing_contact_id += 1


            else:  # Haven't seen header, so this was first row.
                seen_header = True

        # NOTE: we now run this in a engine.begin() context manager, so our
        # parent will commit. Don't commit here!


        logger.debug("Stats: Total rows: %s  Dupes: %s   Missing contact_id: %s",  str(row_count) ,  str(dupes),  str(missing_contact_id) )
        logger.debug("Other integrity exceptions: %s    Other exceptions: %s",  str(other_integrity),  str(other_exceptions) )
        wb.close()
        return { True : "File imported" }

    else:  # Similarity too low 
        wb.close()
        logger.warn("Similarity value of  %s   is below threshold of  %s  so file was not processed ", '{:.2}'.format(min_similarity),   str(MINIMUM_SIMILARITY))  
        return {False : "Similarity to expected column names below threshold"}

