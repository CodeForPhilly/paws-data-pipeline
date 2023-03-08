import re
from flask.globals import current_app
from datetime import datetime, timedelta 
from openpyxl import load_workbook
from jellyfish import jaro_similarity

from config import  engine

import structlog

from api.API_ingest.volgistics import insert_volgistics_people
logger = structlog.get_logger()


from sqlalchemy import  insert,  Table,  Column, MetaData, exc
from sqlalchemy.dialects.postgresql import Insert

from sqlalchemy.orm import sessionmaker


metadata = MetaData()


MINIMUM_SIMILARITY = 0.85  # How good does the table match need to be?

expected_shifts_columns =  {
            'Number' : 'volg_id',
            'Site' : 'site',
            'Place' : None,
            'Assignment' : 'assignment',
            'From date' : 'from_date',
            'To date' : None,
            'From time' :None,
            'To time' : None,
            'Hours' : 'hours',
            'No Call/No Show' : None,
            'Call/Email to miss shift' : None,
            'Absence' : None,
            'Volunteers' : None
            }


def open_volgistics(filename):
    logger.info("Loading '%s' - this is slow",  filename.filename )
    start = datetime.now()
    wb = load_workbook(filename)   #  ,read_only=True should be faster but gets size incorrect 
    end = datetime.now()
    logger.info("Loaded '%s' complete in %d seconds",  filename.filename, (end-start).seconds )
    return wb

def validate_import_vs(workbook, conn):
    """ Validate that the XLSX column names int the file are close enough to expectations that we can trust the data.
        If so, insert the data into the volgisticsshifts table. 
    """

    # logger.info('------ Loading %s ',  filename.filename )
    # wb = load_workbook(filename)   #  ,read_only=True should be faster but gets size incorrect 
    ws = workbook['Service']   # Needs to be 'Service' sheet
    # ws.reset_dimensions()   # Tells openpyxl to ignore what sheet says and check for itself
    ws.calculate_dimension()

    columns = ws.max_column
    if columns > 26:
        # Only 13 actually populated 
        logger.info("Column count > 26; columns after Z not processed")
        columns = 26

    header = [cell.value for cell in ws[1]]

    min_similarity = 1.0
    min_column = None

    for expected, got in zip(expected_shifts_columns.keys(), header):
        jsim = jaro_similarity(expected, got) 
        if jsim < min_similarity :
            min_similarity = jsim
            min_column = expected + ' / ' + got

        
    logger.debug("Minimum similarity:  %s", "{:.2}".format(min_similarity) )
    if min_column:
        logger.debug("On expected/got: %s", str(min_column))


    if  min_similarity >= MINIMUM_SIMILARITY :  # Good enough to trust
        
        vs  = Table("volgisticsshifts", metadata, autoload=True, autoload_with=engine)

        seen_header = False  # Used to skip header row

        # Stats for import
        dupes = 0
        other_integrity = 0
        other_exceptions = 0
        row_count = 0
        missing_volgistics_id = 0


        #TODO: Perform bulk insert as for people_insert


        for row in ws.values:        
            if seen_header: 
                row_count += 1
                if (row_count % 1000 == 0) and (row_count % 5000 != 0):
                    logger.debug("Row: %s", str(row_count) )
                if row_count % 5000 == 0:
                    logger.info("Row: %s", str(row_count) )
                zrow = dict(zip(expected_shifts_columns.values(), row))  
                # zrow is a dict of db_col:value pairs, with at most one key being None (as it overwrote any previous)
                # We need to remove the None item, if it exists
                try:
                    del zrow[None]
                except KeyError:
                    pass 

                #  Cleanup time!  Many older imports have... peculiarities 

                #  End cleanup 

                if zrow['volg_id'] :  # No point in importing if there's nothing to match
                    # Finally ready to insert row into the table
                    # 

                    stmt = Insert(vs).values(zrow)

                    skip_dupes = stmt.on_conflict_do_nothing(
                        constraint='uq_shift'
                       )
                    try:
                        result = conn.execute(skip_dupes)
                    except exc.IntegrityError as e:  # Catch-all for several more specific exceptions
                        if  re.match('duplicate key value', str(e.orig) ):
                            dupes += 1
                            pass
                        else:
                            other_integrity += 1
                            logger.error(e)
                    except Exception as e: 
                        other_exceptions += 1
                        logger.error(e)
                 
                else: # Missing contact_id
                    missing_volgistics_id += 1


            else:  # Haven't seen header, so this was first row.
                seen_header = True

        # NOTE: we now run this in a engine.begin() context manager, so our
        # parent will commit. Don't commit here!


        logger.info("Total rows: %s  Dupes: %s Missing volgistics id: %s",  str(row_count), str(dupes), str(missing_volgistics_id)  )
        logger.info("Other integrity exceptions: %s  Other exceptions: %s",  str(other_exceptions),  str(other_integrity) )
        # workbook.close()
        return { True : "File imported" }
    

def volgistics_people_import(workbook,conn):

    ws = workbook['Master']   # Needs to be 'Service' sheet
    # ws.reset_dimensions()   # Tells openpyxl to ignore what sheet says and check for itself
    ws.calculate_dimension()

    columns = ws.max_column

    #TODO: Validate header row to ensure source cols haven't changed

    Session = sessionmaker(engine)
    session = Session()
    metadata = MetaData()
    volg_table = Table("volgistics", metadata, autoload=True, autoload_with=engine)

    # Worksheet cells are addressed as ws[row][col] with row being 1-based and col being 0-based

    insert_list = []

    # Create a dict from header row so can reference columns by name
    # e.g., r[col['Number']] instead of r[15]
    header = ws[1]
    col = {};
    idx = 0
    for cell in header:
        col[cell.value] = idx
        idx += 1



    time_stamp = datetime.utcnow()

    try:
        for r in ws.iter_rows(min_row=2, max_col=42,values_only=True):
            insert_list.append(
                {
                    "number": r[col['Number']],
                    "last_name": r[col['Last name']],
                    "first_name": r[col['First name']],
                    "middle_name": r[col['Middle name']],
                    "complete_address": r[col['Complete address']],
                    "street_1": r[col['Street 1']],
                    "street_2": r[col['Street 2']],
                    "street_3": r[col['Street 3']],
                    "city": r[col['City']],
                    "state": r[col['State']],
                    "zip": r[col['Zip']],
                    "all_phone_numbers": r[col['All phone numbers']],
                    "home": r[col['Home']],
                    "work": r[col['Work']],
                    "cell": r[col['Cell']],
                    "email": r[col['Email']],
                    "created_date" : time_stamp
                }
            )
    except KeyError as e:
        logger.error("Volgistics source XLSX file 'Master' tab missing expected column (see following)  - cannot import")
        logger.exception(e)



    rows = insert_volgistics_people(insert_list)

    logger.debug('Inserted %d Volgistics people rows', rows)