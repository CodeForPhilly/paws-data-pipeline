from flask.globals import current_app
from datetime import datetime, timedelta 
from openpyxl import load_workbook
from jellyfish import jaro_similarity

from config import  engine
from utils import standardize_phone_number

import structlog

from api.API_ingest.volgistics_db import insert_volgistics_people, insert_volgistics_shifts
logger = structlog.get_logger()


from sqlalchemy import Table, MetaData

from sqlalchemy.orm import sessionmaker


metadata = MetaData()


MINIMUM_SIMILARITY = 0.85  # How good does the table match need to be?

expected_shifts_columns =  {
            'Number' : 'volg_id',
            'Site' : 'site',
            'Place' : None,
            'Assignment' : 'assignment',
            'From date' : 'from_date',
            'To date' : None,
            'From time' :None,
            'To time' : None,
            'Hours' : 'hours',
            'No Call/No Show' : None,
            'Call/Email to miss shift' : None,
            'Absence' : None,
            'Volunteers' : None
            }


def open_volgistics(filename):
    logger.info("Loading '%s' ",  filename.filename )
    start = datetime.now()
    wb = load_workbook(filename, read_only=True)   # read_only=True caused problems (got size wrong) in an earlier version   
                                                   # works fine now and is much faster 
    
    end = datetime.now()
    logger.info("Loaded '%s' complete in %d seconds",  filename.filename, (end-start).seconds )

    try:
        ws = wb['Service']
        wp = wb['Master']

    except Exception as e:
        logger.error("Could not open expected tab in '%s' - not a Volgistics xlsx file?: %s", filename.filename, e )
        return None

    return wb

def validate_import_vs(workbook):
    """ Validate that the XLSX column names int the file are close enough to expectations that we can trust the data.
        If so, insert the data into the volgisticsshifts table. 
    """


    ws = workbook['Service']   # Needs to be 'Service' sheet
    # ws.reset_dimensions()   # Tells openpyxl to ignore what sheet says and check for itself
    ws.calculate_dimension()

    columns = ws.max_column
    if columns > 26:
        # Only 13 actually populated 
        logger.info("Column count > 26; columns after Z not processed")
        columns = 26

    header = [cell.value for cell in ws[1]]

    min_similarity = 1.0
    min_column = None

    for expected, got in zip(expected_shifts_columns.keys(), header):
        jsim = jaro_similarity(expected, got) 
        if jsim < min_similarity :
            min_similarity = jsim
            min_column = expected + ' / ' + got

        
    logger.debug("Minimum similarity:  %s", "{:.2}".format(min_similarity) )
    if min_column:
        logger.debug("On expected/got: %s", str(min_column))


    if  min_similarity >= MINIMUM_SIMILARITY :  # Good enough to trust
        
        seen_header = False  # Used to skip header row

        # Stats for import
        row_count = 0
        missing_volgistics_id = 0

        shifts_rows = [];

        for row in ws.values:        
            if seen_header: 
                row_count += 1
                # if (row_count % 1000 == 0) and (row_count % 5000 != 0):
                #     logger.debug("Row: %s", str(row_count) )
                if row_count % 5000 == 0:
                    logger.info("Row: %s", str(row_count) )
                zrow = dict(zip(expected_shifts_columns.values(), row))  
                # zrow is a dict of db_col:value pairs, with at most one key being None (as it overwrote any previous)
                # We need to remove the None item, if it exists
                try:
                    del zrow[None]
                except KeyError:
                    pass 

                if zrow['volg_id'] :  # No point in importing if there's nothing to match
                    shifts_rows.append(zrow)
                else: # Missing contact_id
                    missing_volgistics_id += 1

            else:  # Haven't seen header, so this was first row.
                seen_header = True

        rows = insert_volgistics_shifts(shifts_rows)

        logger.info("Total rows: %d  Missing volgistics id: %d",  rows, missing_volgistics_id  )
        return { True : "File imported" }
    

def volgistics_people_import(workbook):

    ws = workbook['Master']   # Needs to be 'Master' sheet
    # ws.reset_dimensions()   # Tells openpyxl to ignore what sheet says and check for itself
    ws.calculate_dimension()

    columns = ws.max_column

    #TODO: Validate header row to ensure source cols haven't changed

    Session = sessionmaker(engine)
    session = Session()
    metadata = MetaData()
    volg_table = Table("volgistics", metadata, autoload=True, autoload_with=engine)

    # Worksheet cells are addressed as ws[row][col] with row being 1-based and col being 0-based

    insert_list = []

    # Create a dict from header row so can reference columns by name
    # e.g., r[col['Number']] instead of r[15]
    header = ws[1]
    col = {};
    idx = 0
    for cell in header:
        col[cell.value] = idx
        idx += 1

    # This table has something like 115 columns - not interested in handling each even if empty
    # Get the column numbers of the ones we care about
    col_number = col['Number']
    col_lastname = col['Last name']
    col_firstname =  col['First name']
    col_middlename = col['Middle name']
    col_complete_address = col['Complete address']
    col_street1 = col['Street 1']
    col_street2 = col['Street 2']
    col_street3 = col['Street 3']
    col_city = col['City']
    col_state = col['State']
    col_zip = col['Zip']
    col_all_phones = col['All phone numbers']
    col_home = col['Home']
    col_work = col['Work']
    col_cell = col['Cell']
    col_email = col['Email']
    time_stamp = datetime.utcnow()

    home_phone = standardize_phone_number(r[col_home])
    work_phone = standardize_phone_number(r[col_work])
    cell_phone = standardize_phone_number(r[col_cell])

    try:
        for r in ws.iter_rows(min_row=2, max_col=42,values_only=True):
            insert_list.append(
                {
                    "number": r[col_number],
                    "last_name": r[col_lastname],
                    "first_name": r[col_firstname],
                    "middle_name": r[col_middlename],
                    "complete_address": r[col_complete_address],
                    "street_1": r[col_street1],
                    "street_2": r[col_street2],
                    "street_3": r[col_street3],
                    "city": r[col_city],
                    "state": r[col_state],
                    "zip": r[col_zip],
                    "all_phone_numbers": r[col_all_phones],
                    "home": home_phone,
                    "work": work_phone,
                    "cell": cell_phone,
                    "email": r[col_email],
                    "created_date" : time_stamp
                }
            )
    except KeyError as e:
        logger.error("Volgistics source XLSX file 'Master' tab missing expected column (see following)  - cannot import")
        logger.exception(e)

    except Exception as e:
        logger.error("Unhandled exception preparing Volgistics people records for import")
        logger.exception(e)

    rows = insert_volgistics_people(insert_list)

    logger.debug('Inserted %d Volgistics people rows', rows)